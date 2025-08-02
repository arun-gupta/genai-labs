import requests
from bs4 import BeautifulSoup
import PyPDF2
import io
from docx import Document
import openpyxl
import markdown
from typing import Optional, Tuple
import re


class InputProcessor:
    """Service for processing different types of input for summarization."""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
    
    def extract_text_from_url(self, url: str) -> Tuple[str, str]:
        """Extract text content from a URL."""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Get title
            title = soup.find('title')
            title_text = title.get_text().strip() if title else "No title"
            
            # Get main content
            # Try to find main content areas
            main_content = ""
            
            # Look for common content containers
            content_selectors = [
                'main', 'article', '.content', '.post-content', '.entry-content',
                '.article-content', '.main-content', '#content', '#main'
            ]
            
            for selector in content_selectors:
                content = soup.select_one(selector)
                if content:
                    main_content = content.get_text(separator=' ', strip=True)
                    break
            
            # If no main content found, get body text
            if not main_content:
                main_content = soup.get_text(separator=' ', strip=True)
            
            # Clean up the text
            main_content = self._clean_text(main_content)
            
            # Combine title and content
            full_text = f"Title: {title_text}\n\n{main_content}"
            
            return full_text, title_text
            
        except Exception as e:
            raise ValueError(f"Failed to extract content from URL: {str(e)}")
    
    def extract_text_from_file(self, file_content: bytes, file_type: str) -> str:
        """Extract text content from uploaded file."""
        try:
            if file_type.lower() == 'txt':
                return file_content.decode('utf-8')
            
            elif file_type.lower() == 'pdf':
                return self._extract_from_pdf(file_content)
            
            elif file_type.lower() in ['docx', 'doc']:
                return self._extract_from_docx(file_content)
            
            elif file_type.lower() in ['xlsx', 'xls']:
                return self._extract_from_excel(file_content)
            
            elif file_type.lower() in ['md', 'markdown']:
                return self._extract_from_markdown(file_content)
            
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
                
        except Exception as e:
            raise ValueError(f"Failed to extract content from file: {str(e)}")
    
    def _extract_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file."""
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Failed to extract text from PDF: {str(e)}")
    
    def _extract_from_docx(self, file_content: bytes) -> str:
        """Extract text from DOCX file."""
        try:
            doc_file = io.BytesIO(file_content)
            doc = Document(doc_file)
            
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Failed to extract text from DOCX: {str(e)}")
    
    def _extract_from_excel(self, file_content: bytes) -> str:
        """Extract text from Excel file."""
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Failed to extract text from Excel: {str(e)}")
    
    def _extract_from_markdown(self, file_content: bytes) -> str:
        """Extract text from Markdown file."""
        try:
            md_text = file_content.decode('utf-8')
            # Convert markdown to plain text
            html = markdown.markdown(md_text)
            soup = BeautifulSoup(html, 'html.parser')
            return soup.get_text(separator=' ', strip=True)
        except Exception as e:
            raise ValueError(f"Failed to extract text from Markdown: {str(e)}")
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize text content."""
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text)
        
        # Remove excessive newlines
        text = re.sub(r'\n\s*\n\s*\n', '\n\n', text)
        
        # Remove special characters that might interfere with summarization
        text = re.sub(r'[^\w\s\.\,\!\?\;\:\-\(\)\[\]\{\}]', '', text)
        
        # Truncate text to avoid context length issues (roughly 3000 tokens)
        words = text.split()
        if len(words) > 3500:  # Conservative estimate: ~3500 words = ~3000 tokens
            text = ' '.join(words[:3500]) + '... [Content truncated for length]'
        
        return text.strip()
    
    def validate_input(self, text: Optional[str] = None, url: Optional[str] = None, 
                      file_content: Optional[str] = None, file_type: Optional[str] = None) -> str:
        """Validate and process input, returning the text to summarize."""
        input_count = sum(1 for x in [text, url, file_content] if x is not None)
        
        if input_count == 0:
            raise ValueError("No input provided. Please provide text, URL, or file content.")
        
        if input_count > 1:
            raise ValueError("Please provide only one input type: text, URL, or file content.")
        
        if text:
            return self._clean_text(text)
        
        elif url:
            extracted_text, title = self.extract_text_from_url(url)
            return extracted_text
        
        elif file_content and file_type:
            # Convert file_content from base64 or string to bytes if needed
            if isinstance(file_content, str):
                # Assume it's base64 encoded
                import base64
                try:
                    file_bytes = base64.b64decode(file_content)
                except:
                    # If not base64, treat as plain text
                    return self._clean_text(file_content)
            else:
                file_bytes = file_content
            
            return self.extract_text_from_file(file_bytes, file_type)
        
        else:
            raise ValueError("Invalid input combination.")


# Global instance
input_processor = InputProcessor() 